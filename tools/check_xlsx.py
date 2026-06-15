import openpyxl, sys

path = sys.argv[1]
wb = openpyxl.load_workbook(path)
print("Sheets:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    print(f"\n--- {name} ({len(rows)} rows) ---")
    for r in rows[:5]:
        print(r)
    if len(rows) > 5:
        print(f"  ... ({len(rows)-5} more)")
